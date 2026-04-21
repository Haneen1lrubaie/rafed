from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify
from werkzeug.security import generate_password_hash, check_password_hash
import os, io, random, smtplib, secrets
from datetime import date, datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from functools import wraps
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
try:
    import arabic_reshaper
    from bidi.algorithm import get_display
    _ARABIC_SUPPORT = True
except ImportError:
    _ARABIC_SUPPORT = False
import psycopg2
from psycopg2.extras import RealDictCursor
import cloudinary
import cloudinary.uploader

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'rafed-dev-key-change-in-production')

cloudinary.config(
    cloud_name = os.environ.get('CLOUDINARY_CLOUD_NAME', ''),
    api_key    = os.environ.get('CLOUDINARY_API_KEY', ''),
    api_secret = os.environ.get('CLOUDINARY_API_SECRET', ''),
    secure     = True
)

MAIL_HOST     = os.environ.get('MAIL_HOST',     'smtp.gmail.com')
MAIL_PORT     = int(os.environ.get('MAIL_PORT', '587'))
MAIL_USER     = os.environ.get('MAIL_USER',     '')
MAIL_PASSWORD = os.environ.get('MAIL_PASSWORD', '')
MAIL_FROM     = os.environ.get('MAIL_FROM',     'noreply@rafed.sa')
DATABASE_URL  = os.environ.get('DATABASE_URL',  '')
ALLOWED_EXTENSIONS = {'pdf', 'pptx'}

# ─────────────────────────────────────────
# DATABASE
# ─────────────────────────────────────────
def get_db():
    return psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)

def init_db():
    conn = get_db(); c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        id SERIAL PRIMARY KEY, name TEXT NOT NULL, email TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL, role TEXT NOT NULL DEFAULT 'trainee',
        group_id INTEGER, is_active INTEGER DEFAULT 1,
        email_verified INTEGER DEFAULT 0, created_at TIMESTAMP DEFAULT NOW())''')
    # ✅ IMPROVEMENT 2: groups no longer have a single supervisor_id
    c.execute('''CREATE TABLE IF NOT EXISTS groups (
        id SERIAL PRIMARY KEY, name TEXT NOT NULL,
        start_date TEXT, created_at TIMESTAMP DEFAULT NOW())''')
    # ✅ IMPROVEMENT 2: new table for multiple supervisors per group
    c.execute('''CREATE TABLE IF NOT EXISTS group_supervisors (
        id SERIAL PRIMARY KEY,
        group_id INTEGER NOT NULL REFERENCES groups(id) ON DELETE CASCADE,
        supervisor_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
        assigned_at TIMESTAMP DEFAULT NOW(),
        UNIQUE(group_id, supervisor_id))''')
    c.execute('''CREATE TABLE IF NOT EXISTS attendance (
        id SERIAL PRIMARY KEY, user_id INTEGER NOT NULL,
        date TEXT NOT NULL, status TEXT DEFAULT 'present',
        recorded_by INTEGER, created_at TIMESTAMP DEFAULT NOW(),
        UNIQUE(user_id, date))''')
    c.execute('''CREATE TABLE IF NOT EXISTS initiatives (
        id SERIAL PRIMARY KEY, user_id INTEGER NOT NULL,
        filename TEXT NOT NULL, original_name TEXT NOT NULL,
        file_url TEXT, status TEXT DEFAULT 'pending',
        feedback TEXT, reviewed_by INTEGER, reviewed_at TEXT,
        uploaded_at TIMESTAMP DEFAULT NOW())''')
    # ✅ IMPROVEMENT 4: invite codes only for supervisors
    c.execute('''CREATE TABLE IF NOT EXISTS invite_codes (
        id SERIAL PRIMARY KEY, code TEXT UNIQUE NOT NULL,
        role TEXT NOT NULL DEFAULT 'supervisor',
        used INTEGER DEFAULT 0,
        created_by INTEGER, created_at TIMESTAMP DEFAULT NOW())''')
    c.execute('''CREATE TABLE IF NOT EXISTS otp_codes (
        id SERIAL PRIMARY KEY, email TEXT NOT NULL,
        code TEXT NOT NULL, expires_at TIMESTAMP NOT NULL,
        used INTEGER DEFAULT 0, created_at TIMESTAMP DEFAULT NOW())''')
    c.execute('''CREATE TABLE IF NOT EXISTS password_resets (
        id SERIAL PRIMARY KEY, email TEXT NOT NULL,
        token TEXT NOT NULL, expires_at TIMESTAMP NOT NULL,
        used INTEGER DEFAULT 0, created_at TIMESTAMP DEFAULT NOW())''')
    c.execute("SELECT id FROM users WHERE role='admin'")
    if not c.fetchone():
        c.execute(
            "INSERT INTO users (name,email,password,role,email_verified) VALUES (%s,%s,%s,%s,1)",
            ('مدير النظام', 'rafed26program@gmail.com',
             generate_password_hash('Rafed@2026'), 'admin'))
    conn.commit(); conn.close()

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def upload_to_cloudinary(file, filename):
    try:
        ext = filename.rsplit('.', 1)[1].lower()
        result = cloudinary.uploader.upload(
            file, resource_type='raw', folder='rafed/initiatives',
            public_id=f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{secrets.token_hex(4)}",
            format=ext, use_filename=False, access_mode='public')
        return result['public_id'], result['secure_url']
    except Exception as e:
        print(f"Cloudinary error: {e}")
        return None, None

def delete_from_cloudinary(public_id):
    try:
        cloudinary.uploader.destroy(public_id, resource_type='raw')
    except Exception as e:
        print(f"Cloudinary delete error: {e}")

# ─────────────────────────────────────────
# EMAIL & OTP
# ─────────────────────────────────────────
def generate_otp():
    return str(random.randint(100000, 999999))

def send_email(to_email, subject, html_body):
    if not MAIL_USER or not MAIL_PASSWORD:
        print(f"\n{'='*40}\n[DEV] Email to {to_email}\nSubject: {subject}\n{'='*40}\n")
        return True, ''
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = MAIL_FROM
        msg['To']      = to_email
        msg.attach(MIMEText(html_body, 'html', 'utf-8'))
        with smtplib.SMTP(MAIL_HOST, MAIL_PORT, timeout=10) as s:
            s.ehlo(); s.starttls()
            s.login(MAIL_USER, MAIL_PASSWORD)
            s.sendmail(MAIL_FROM, to_email, msg.as_string())
        return True, ''
    except smtplib.SMTPAuthenticationError:
        return False, 'فشل المصادقة — تحقق من إعدادات Gmail'
    except Exception as e:
        return False, str(e)

def send_otp_email(to_email, otp_code, name=''):
    html = f"""
    <div dir="rtl" style="font-family:Arial,sans-serif;max-width:480px;margin:auto;
         border:1px solid #e0e0e0;border-radius:12px;overflow:hidden;">
      <div style="background:#0F6E56;padding:28px;text-align:center;">
        <h1 style="color:white;margin:0;font-size:2rem;">رفد</h1>
      </div>
      <div style="padding:32px;background:#fff;">
        <h2 style="color:#2C2C2A;margin-top:0;">مرحباً {name}</h2>
        <p style="color:#5F5E5A;">أدخل رمز التحقق التالي لإتمام تسجيلك:</p>
        <div style="background:#E1F5EE;border-radius:12px;padding:24px;text-align:center;margin:24px 0;">
          <span style="font-size:2.5rem;font-weight:bold;letter-spacing:12px;color:#0F6E56;">{otp_code}</span>
        </div>
        <p style="color:#854F0B;font-size:0.9rem;">⚠️ صالح لمدة <strong>10 دقائق</strong> فقط.</p>
      </div>
      <div style="background:#F1EFE8;padding:16px;text-align:center;">
        <p style="color:#888;font-size:0.8rem;margin:0;">منصة رفد</p>
      </div>
    </div>"""
    return send_email(to_email, 'رمز التحقق — منصة رفد', html)

def send_reset_email(to_email, reset_link, name=''):
    html = f"""
    <div dir="rtl" style="font-family:Arial,sans-serif;max-width:480px;margin:auto;
         border:1px solid #e0e0e0;border-radius:12px;overflow:hidden;">
      <div style="background:#0F6E56;padding:28px;text-align:center;">
        <h1 style="color:white;margin:0;font-size:2rem;">رفد</h1>
      </div>
      <div style="padding:32px;background:#fff;">
        <h2 style="color:#2C2C2A;margin-top:0;">مرحباً {name}</h2>
        <p style="color:#5F5E5A;">طلبت إعادة تعيين كلمة المرور. اضغط على الزر أدناه:</p>
        <div style="text-align:center;margin:24px 0;">
          <a href="{reset_link}" style="background:#0F6E56;color:white;padding:14px 28px;
             border-radius:8px;text-decoration:none;font-size:1rem;font-weight:bold;">
            إعادة تعيين كلمة المرور
          </a>
        </div>
        <p style="color:#854F0B;font-size:0.9rem;">⚠️ الرابط صالح لمدة <strong>30 دقيقة</strong> فقط.</p>
      </div>
      <div style="background:#F1EFE8;padding:16px;text-align:center;">
        <p style="color:#888;font-size:0.8rem;margin:0;">منصة رفد</p>
      </div>
    </div>"""
    return send_email(to_email, 'إعادة تعيين كلمة المرور — رفد', html)

def save_otp(email, code):
    expires = datetime.now() + timedelta(minutes=10)
    conn = get_db(); c = conn.cursor()
    c.execute("UPDATE otp_codes SET used=1 WHERE email=%s AND used=0", (email,))
    c.execute("INSERT INTO otp_codes (email,code,expires_at) VALUES (%s,%s,%s)", (email, code, expires))
    conn.commit(); conn.close()

def verify_otp(email, code):
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM otp_codes WHERE email=%s AND code=%s AND used=0 ORDER BY id DESC LIMIT 1",
              (email, code))
    row = c.fetchone()
    if not row or datetime.now() > row['expires_at']:
        conn.close(); return False
    c.execute("UPDATE otp_codes SET used=1 WHERE id=%s", (row['id'],))
    conn.commit(); conn.close()
    return True

# ─────────────────────────────────────────
# AUTH DECORATORS
# ─────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('يجب تسجيل الدخول أولاً', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if session.get('role') not in roles:
                flash('ليس لديك صلاحية للوصول لهذه الصفحة', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated
    return decorator

# ─────────────────────────────────────────
# PUBLIC ROUTES
# ─────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        email    = request.form['email'].strip().lower()
        password = request.form['password']
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT * FROM users WHERE email=%s AND is_active=1", (email,))
        user = c.fetchone(); conn.close()
        if user and check_password_hash(user['password'], password):
            session.update({'user_id': user['id'], 'name': user['name'],
                            'role': user['role'], 'group_id': user['group_id']})
            return redirect(url_for('dashboard'))
        flash('البريد الإلكتروني أو كلمة المرور غير صحيحة', 'danger')
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        name     = request.form['name'].strip()
        email    = request.form['email'].strip().lower()
        password = request.form['password']
        code     = request.form.get('invite_code', '').strip()
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT id FROM users WHERE email=%s", (email,))
        if c.fetchone():
            conn.close()
            flash('البريد الإلكتروني مسجّل مسبقاً', 'danger')
            return render_template('register.html')
        # ✅ IMPROVEMENT 4: only supervisor role allowed via invite code
        role = 'trainee'
        if code:
            c.execute("SELECT * FROM invite_codes WHERE code=%s AND used=0 AND role='supervisor'", (code,))
            inv = c.fetchone()
            if not inv:
                conn.close()
                flash('كود الدعوة غير صحيح أو مستخدم', 'danger')
                return render_template('register.html')
            role = 'supervisor'
        conn.close()
        otp = generate_otp()
        ok, err = send_otp_email(email, otp, name)
        if not ok:
            flash(f'تعذّر إرسال رمز التحقق: {err}', 'danger')
            return render_template('register.html')
        save_otp(email, otp)
        session['pending_reg'] = {
            'name': name, 'email': email,
            'password': generate_password_hash(password),
            'role': role, 'invite_code': code
        }
        flash('تم إرسال رمز التحقق إلى بريدك الإلكتروني', 'info')
        return redirect(url_for('verify_otp_page'))
    return render_template('register.html')

@app.route('/verify-otp', methods=['GET', 'POST'])
def verify_otp_page():
    if 'pending_reg' not in session:
        return redirect(url_for('register'))
    if request.method == 'POST':
        entered = request.form['otp'].strip()
        email   = session['pending_reg']['email']
        if not verify_otp(email, entered):
            flash('رمز التحقق غير صحيح أو منتهي الصلاحية', 'danger')
            return render_template('verify_otp.html', email=email)
        reg = session.pop('pending_reg')
        conn = get_db(); c = conn.cursor()
        try:
            c.execute(
                "INSERT INTO users (name,email,password,role,email_verified) VALUES (%s,%s,%s,%s,1)",
                (reg['name'], reg['email'], reg['password'], reg['role']))
            if reg.get('invite_code'):
                c.execute("UPDATE invite_codes SET used=1 WHERE code=%s", (reg['invite_code'],))
            conn.commit()
        except Exception:
            conn.close()
            flash('حدث خطأ أثناء إنشاء الحساب، حاول مجدداً', 'danger')
            return redirect(url_for('register'))
        conn.close()
        flash('تم إنشاء حسابك بنجاح! يمكنك تسجيل الدخول الآن', 'success')
        return redirect(url_for('login'))
    return render_template('verify_otp.html', email=session['pending_reg']['email'])

@app.route('/api/resend-otp', methods=['POST'])
def resend_otp():
    if 'pending_reg' not in session:
        return jsonify({'ok': False, 'msg': 'انتهت جلسة التسجيل'})
    reg = session['pending_reg']
    otp = generate_otp()
    ok, err = send_otp_email(reg['email'], otp, reg['name'])
    if not ok:
        return jsonify({'ok': False, 'msg': err})
    save_otp(reg['email'], otp)
    return jsonify({'ok': True, 'msg': 'تم إعادة إرسال رمز التحقق'})

@app.route('/api/check-email', methods=['POST'])
def check_email():
    email = (request.json or {}).get('email', '').strip().lower()
    if not email:
        return jsonify({'available': False, 'msg': 'أدخل البريد الإلكتروني'})
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT id FROM users WHERE email=%s", (email,))
    exists = c.fetchone(); conn.close()
    if exists:
        return jsonify({'available': False, 'msg': 'هذا البريد مسجّل مسبقاً'})
    return jsonify({'available': True, 'msg': 'البريد متاح'})

@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form['email'].strip().lower()
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT * FROM users WHERE email=%s AND is_active=1", (email,))
        user = c.fetchone()
        if user:
            token   = secrets.token_urlsafe(32)
            expires = datetime.now() + timedelta(minutes=30)
            c.execute("UPDATE password_resets SET used=1 WHERE email=%s", (email,))
            c.execute("INSERT INTO password_resets (email,token,expires_at) VALUES (%s,%s,%s)",
                      (email, token, expires))
            conn.commit()
            reset_link = url_for('reset_password', token=token, _external=True)
            send_reset_email(email, reset_link, user['name'])
        conn.close()
        flash('إذا كان البريد مسجّلاً ستصلك رسالة لإعادة تعيين كلمة المرور', 'info')
        return redirect(url_for('login'))
    return render_template('forgot_password.html')

@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM password_resets WHERE token=%s AND used=0 ORDER BY id DESC LIMIT 1", (token,))
    row = c.fetchone()
    if not row or datetime.now() > row['expires_at']:
        conn.close()
        flash('رابط إعادة التعيين منتهي الصلاحية أو غير صحيح', 'danger')
        return redirect(url_for('forgot_password'))
    if request.method == 'POST':
        new_pass = request.form['password']
        if len(new_pass) < 6:
            flash('كلمة المرور يجب أن تكون 6 أحرف على الأقل', 'danger')
            return render_template('reset_password.html', token=token)
        c.execute("UPDATE users SET password=%s WHERE email=%s",
                  (generate_password_hash(new_pass), row['email']))
        c.execute("UPDATE password_resets SET used=1 WHERE token=%s", (token,))
        conn.commit(); conn.close()
        flash('تم تغيير كلمة المرور بنجاح', 'success')
        return redirect(url_for('login'))
    conn.close()
    return render_template('reset_password.html', token=token)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

# ─────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────
@app.route('/dashboard')
@login_required
def dashboard():
    role = session['role']
    if role == 'admin':      return redirect(url_for('admin_dashboard'))
    if role == 'supervisor': return redirect(url_for('supervisor_dashboard'))
    return redirect(url_for('trainee_dashboard'))

# ─────────────────────────────────────────
# TRAINEE
# ─────────────────────────────────────────
@app.route('/trainee')
@login_required
@role_required('trainee')
def trainee_dashboard():
    uid = session['user_id']; today = str(date.today())
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT id FROM attendance WHERE user_id=%s AND date=%s", (uid, today))
    checked_today = c.fetchone()
    c.execute("SELECT * FROM attendance WHERE user_id=%s ORDER BY date DESC", (uid,))
    attendance = c.fetchall()
    c.execute("SELECT * FROM initiatives WHERE user_id=%s ORDER BY uploaded_at DESC", (uid,))
    initiatives = c.fetchall()
    c.execute("SELECT u.*, g.name as group_name, g.start_date FROM users u "
              "LEFT JOIN groups g ON u.group_id=g.id WHERE u.id=%s", (uid,))
    user = c.fetchone()
    conn.close()
    total_days   = len(attendance)
    present_days = sum(1 for a in attendance if a['status'] == 'present')
    absent_days  = total_days - present_days
    return render_template('trainee_dashboard.html',
                           checked_today=checked_today, attendance=attendance,
                           initiatives=initiatives, today=today, user=user,
                           total_days=total_days, present_days=present_days,
                           absent_days=absent_days)

# ✅ IMPROVEMENT 3: removed trainee self-checkin — attendance only by supervisor
# ✅ IMPROVEMENT 5: trainee can delete pending initiative
@app.route('/trainee/initiative/<int:init_id>/delete', methods=['POST'])
@login_required
@role_required('trainee')
def trainee_delete_initiative(init_id):
    uid = session['user_id']
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM initiatives WHERE id=%s AND user_id=%s", (init_id, uid))
    ini = c.fetchone()
    if not ini:
        conn.close()
        flash('الملف غير موجود', 'danger')
        return redirect(url_for('trainee_dashboard'))
    if ini['status'] != 'pending':
        conn.close()
        flash('لا يمكن حذف ملف تم تقييمه', 'danger')
        return redirect(url_for('trainee_dashboard'))
    # Delete from Cloudinary
    if ini['filename']:
        delete_from_cloudinary(ini['filename'])
    c.execute("DELETE FROM initiatives WHERE id=%s", (init_id,))
    conn.commit(); conn.close()
    flash('تم حذف الملف بنجاح', 'success')
    return redirect(url_for('trainee_dashboard'))

@app.route('/trainee/upload', methods=['POST'])
@login_required
@role_required('trainee')
def trainee_upload():
    if 'file' not in request.files or request.files['file'].filename == '':
        flash('لم يتم اختيار ملف', 'danger')
        return redirect(url_for('trainee_dashboard'))
    file = request.files['file']
    if not allowed_file(file.filename):
        flash('نوع الملف غير مسموح. يُقبل PDF و PPTX فقط', 'danger')
        return redirect(url_for('trainee_dashboard'))
    original = file.filename
    public_id, file_url = upload_to_cloudinary(file, original)
    if not file_url:
        flash('فشل رفع الملف، حاول مجدداً', 'danger')
        return redirect(url_for('trainee_dashboard'))
    conn = get_db(); c = conn.cursor()
    c.execute("INSERT INTO initiatives (user_id,filename,original_name,file_url) VALUES (%s,%s,%s,%s)",
              (session['user_id'], public_id, original, file_url))
    conn.commit(); conn.close()
    flash('تم رفع الملف بنجاح', 'success')
    return redirect(url_for('trainee_dashboard'))

# ─────────────────────────────────────────
# SUPERVISOR
# ─────────────────────────────────────────
@app.route('/supervisor')
@login_required
@role_required('supervisor')
def supervisor_dashboard():
    uid = session['user_id']
    conn = get_db(); c = conn.cursor()
    # ✅ IMPROVEMENT 2: supervisor can manage multiple groups
    c.execute("""SELECT g.* FROM groups g
                 JOIN group_supervisors gs ON g.id=gs.group_id
                 WHERE gs.supervisor_id=%s""", (uid,))
    my_groups = c.fetchall()
    # Use selected group or first group
    selected_gid = request.args.get('group_id', type=int)
    group = None
    if my_groups:
        if selected_gid:
            group = next((g for g in my_groups if g['id'] == selected_gid), my_groups[0])
        else:
            group = my_groups[0]

    trainees = attendance_today = initiatives = attendance_full = []
    present_ids = set()
    if group:
        c.execute("SELECT * FROM users WHERE group_id=%s AND role='trainee' AND is_active=1", (group['id'],))
        trainees = c.fetchall()
        today = str(date.today())
        c.execute("SELECT a.*, u.name FROM attendance a JOIN users u ON a.user_id=u.id "
                  "WHERE u.group_id=%s AND a.date=%s", (group['id'], today))
        attendance_today = c.fetchall()
        present_ids = {a['user_id'] for a in attendance_today}
        c.execute("SELECT a.*, u.name as trainee_name FROM attendance a "
                  "JOIN users u ON a.user_id=u.id "
                  "WHERE u.group_id=%s ORDER BY a.date DESC, u.name", (group['id'],))
        attendance_full = c.fetchall()
        c.execute("SELECT i.*, u.name FROM initiatives i JOIN users u ON i.user_id=u.id "
                  "WHERE u.group_id=%s ORDER BY i.uploaded_at DESC", (group['id'],))
        initiatives = c.fetchall()
    conn.close()
    return render_template('supervisor_dashboard.html',
                           my_groups=my_groups, group=group,
                           trainees=trainees, present_ids=present_ids,
                           attendance_today=attendance_today,
                           attendance_full=attendance_full,
                           initiatives=initiatives, today=str(date.today()))

@app.route('/supervisor/attendance', methods=['POST'])
@login_required
@role_required('supervisor')
def supervisor_attendance():
    user_id = request.form['user_id']; att_date = request.form['date']
    status  = request.form.get('status', 'present')
    conn = get_db(); c = conn.cursor()
    try:
        c.execute("INSERT INTO attendance (user_id,date,status,recorded_by) VALUES (%s,%s,%s,%s)",
                  (user_id, att_date, status, session['user_id']))
        conn.commit(); flash('تم تسجيل الحضور', 'success')
    except psycopg2.errors.UniqueViolation:
        conn.rollback()
        c.execute("UPDATE attendance SET status=%s, recorded_by=%s WHERE user_id=%s AND date=%s",
                  (status, session['user_id'], user_id, att_date))
        conn.commit(); flash('تم تحديث سجل الحضور', 'info')
    conn.close()
    gid = request.form.get('group_id', '')
    return redirect(url_for('supervisor_dashboard', group_id=gid))

@app.route('/supervisor/review/<int:init_id>', methods=['POST'])
@login_required
@role_required('supervisor', 'admin')
def review_initiative(init_id):
    conn = get_db(); c = conn.cursor()
    c.execute("UPDATE initiatives SET status=%s,feedback=%s,reviewed_by=%s,reviewed_at=%s WHERE id=%s",
              (request.form['status'], request.form.get('feedback', ''),
               session['user_id'], datetime.now().isoformat(), init_id))
    conn.commit(); conn.close()
    flash('تم تحديث حالة المبادرة', 'success')
    return redirect(request.referrer or url_for('supervisor_dashboard'))

@app.route('/supervisor/export')
@login_required
@role_required('supervisor')
def supervisor_export():
    fmt = request.args.get('format', 'excel')
    gid = request.args.get('group_id', type=int)
    conn = get_db(); c = conn.cursor()
    if gid:
        c.execute("SELECT * FROM groups WHERE id=%s", (gid,))
        group = c.fetchone()
    else:
        c.execute("""SELECT g.* FROM groups g
                     JOIN group_supervisors gs ON g.id=gs.group_id
                     WHERE gs.supervisor_id=%s LIMIT 1""", (session['user_id'],))
        group = c.fetchone()
    if not group:
        conn.close(); flash('لا توجد مجموعة', 'warning')
        return redirect(url_for('supervisor_dashboard'))
    c.execute("SELECT u.name, a.date, a.status FROM attendance a "
              "JOIN users u ON a.user_id=u.id WHERE u.group_id=%s ORDER BY a.date DESC, u.name",
              (group['id'],))
    records = c.fetchall(); conn.close()
    return export_attendance(records, group['name'], fmt)

# ─────────────────────────────────────────
# EXPORT HELPER
# ─────────────────────────────────────────
def export_attendance(records, group_name, fmt):
    rows = [['الاسم', 'التاريخ', 'الحالة']]
    for r in records:
        rows.append([r['name'], str(r['date']), 'حاضر' if r['status'] == 'present' else 'غائب'])
    export_date = date.today().strftime('%Y-%m-%d')

    if fmt == 'excel':
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'سجل الحضور'
        center = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:C1'); ws['A1'] = 'منصة رفد'
        ws['A1'].font      = openpyxl.styles.Font(bold=True, size=14, color='FFFFFF')
        ws['A1'].fill      = openpyxl.styles.PatternFill('solid', fgColor='0F6E56')
        ws['A1'].alignment = center; ws.row_dimensions[1].height = 28
        ws.merge_cells('A2:C2'); ws['A2'] = f'المجموعة: {group_name}'
        ws['A2'].font      = openpyxl.styles.Font(bold=True, size=12, color='0F6E56')
        ws['A2'].fill      = openpyxl.styles.PatternFill('solid', fgColor='E1F5EE')
        ws['A2'].alignment = center; ws.row_dimensions[2].height = 22
        ws.merge_cells('A3:C3'); ws['A3'] = f'تاريخ التصدير: {export_date}'
        ws['A3'].font      = openpyxl.styles.Font(size=10, color='5F5E5A')
        ws['A3'].alignment = center; ws.row_dimensions[3].height = 18
        hf = openpyxl.styles.Font(bold=True, color='FFFFFF')
        hfill = openpyxl.styles.PatternFill('solid', fgColor='085041')
        for ci, v in enumerate(rows[0], 1):
            cell = ws.cell(row=4, column=ci, value=v)
            cell.font = hf; cell.fill = hfill; cell.alignment = center
        pf = openpyxl.styles.PatternFill('solid', fgColor='EAF3DE')
        af = openpyxl.styles.PatternFill('solid', fgColor='FCEBEB')
        for ri, row in enumerate(rows[1:], 5):
            for ci, v in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.alignment = center
                if ci == 3:
                    cell.fill = pf if v == 'حاضر' else af
                    cell.font = openpyxl.styles.Font(bold=True,
                        color='3B6D11' if v == 'حاضر' else 'A32D2D')
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 16
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name=f'حضور_{group_name}_{export_date}.xlsx')

    # ── Arabic PDF with FreeSerif font ──
    try:
        _font_dir = os.path.join(os.path.dirname(__file__), 'static')
        pdfmetrics.registerFont(TTFont('ArabicFont', os.path.join(_font_dir, 'FreeSerif.ttf')))
        pdfmetrics.registerFont(TTFont('ArabicFont-Bold', os.path.join(_font_dir, 'FreeSerifBold.ttf')))
        FONT      = 'ArabicFont'
        FONT_BOLD = 'ArabicFont-Bold'
    except Exception:
        # Fallback to system fonts
        try:
            pdfmetrics.registerFont(TTFont('ArabicFont', '/usr/share/fonts/truetype/freefont/FreeSerif.ttf'))
            pdfmetrics.registerFont(TTFont('ArabicFont-Bold', '/usr/share/fonts/truetype/freefont/FreeSerifBold.ttf'))
            FONT      = 'ArabicFont'
            FONT_BOLD = 'ArabicFont-Bold'
        except Exception:
            FONT      = 'Helvetica'
            FONT_BOLD = 'Helvetica-Bold'

    def ar(text):
        if _ARABIC_SUPPORT and any('\u0600' <= c <= '\u06FF' for c in str(text)):
            try:
                reshaped = arabic_reshaper.reshape(str(text))
                return get_display(reshaped)
            except Exception:
                return str(text)
        return str(text)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    ts = ParagraphStyle('t', fontName=FONT_BOLD, fontSize=16,
                        textColor=colors.HexColor('#0F6E56'), alignment=1, spaceAfter=4)
    ss = ParagraphStyle('s', fontName=FONT_BOLD, fontSize=12,
                        textColor=colors.HexColor('#085041'), alignment=1, spaceAfter=2)
    ds = ParagraphStyle('d', fontName=FONT, fontSize=9,
                        textColor=colors.HexColor('#5F5E5A'), alignment=1, spaceAfter=12)

    story = [
        Paragraph(ar('منصة رفد'), ts),
        Paragraph(ar(f'المجموعة: {group_name}'), ss),
        Paragraph(f'Export Date: {export_date}', ds),
        Spacer(1, 8),
    ]

    # Build table with Arabic support
    display_rows = [[ar('الحالة'), ar('التاريخ'), ar('الاسم')]]
    for r in rows[1:]:
        status_ar = ar('حاضر') if r[2] == 'حاضر' else ar('غائب')
        display_rows.append([status_ar, r[1], ar(r[0])])

    t = Table(display_rows, colWidths=[90, 110, 250])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0),(-1,0),  colors.HexColor('#085041')),
        ('TEXTCOLOR',  (0,0),(-1,0),  colors.white),
        ('FONTNAME',   (0,0),(-1,0),  FONT_BOLD),
        ('FONTNAME',   (0,1),(-1,-1), FONT),
        ('FONTSIZE',   (0,0),(-1,-1), 10),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#F1EFE8')]),
        ('BOX',  (0,0),(-1,-1), 0.8, colors.HexColor('#0F6E56')),
        ('GRID', (0,0),(-1,-1), 0.3, colors.HexColor('#D3D1C7')),
        ('ALIGN',  (0,0),(-1,-1), 'CENTER'),
        ('VALIGN', (0,0),(-1,-1), 'MIDDLE'),
        ('TOPPADDING',    (0,0),(-1,-1), 7),
        ('BOTTOMPADDING', (0,0),(-1,-1), 7),
    ]))
    # Color present/absent rows
    for i, r in enumerate(rows[1:], 1):
        if r[2] == 'حاضر':
            t.setStyle(TableStyle([
                ('TEXTCOLOR',  (0,i),(0,i), colors.HexColor('#3B6D11')),
                ('BACKGROUND', (0,i),(0,i), colors.HexColor('#EAF3DE')),
                ('FONTNAME',   (0,i),(0,i), FONT_BOLD),
            ]))
        else:
            t.setStyle(TableStyle([
                ('TEXTCOLOR',  (0,i),(0,i), colors.HexColor('#A32D2D')),
                ('BACKGROUND', (0,i),(0,i), colors.HexColor('#FCEBEB')),
                ('FONTNAME',   (0,i),(0,i), FONT_BOLD),
            ]))
    story.append(t)

    total   = len(rows) - 1
    present = sum(1 for r in rows[1:] if r[2] == 'حاضر')
    absent  = total - present
    story.append(Spacer(1, 12))
    s = Table(
        [[ar('الإجمالي'), ar('حاضر'), ar('غائب')],
         [str(total), str(present), str(absent)]],
        colWidths=[150, 150, 150]
    )
    s.setStyle(TableStyle([
        ('BACKGROUND', (0,0),(-1,0), colors.HexColor('#2C2C2A')),
        ('TEXTCOLOR',  (0,0),(-1,0), colors.white),
        ('FONTNAME',   (0,0),(-1,-1), FONT_BOLD),
        ('FONTSIZE',   (0,0),(-1,-1), 10),
        ('BACKGROUND', (1,1),(1,1), colors.HexColor('#EAF3DE')),
        ('BACKGROUND', (2,1),(2,1), colors.HexColor('#FCEBEB')),
        ('TEXTCOLOR',  (1,1),(1,1), colors.HexColor('#3B6D11')),
        ('TEXTCOLOR',  (2,1),(2,1), colors.HexColor('#A32D2D')),
        ('BOX',  (0,0),(-1,-1), 0.5, colors.grey),
        ('GRID', (0,0),(-1,-1), 0.3, colors.grey),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6),
    ]))
    story.append(s)
    doc.build(story); buf.seek(0)
    return send_file(buf, mimetype='application/pdf', as_attachment=True,
                     download_name=f'حضور_{group_name}_{export_date}.pdf')

# ─────────────────────────────────────────
# ADMIN
# ─────────────────────────────────────────
@app.route('/admin')
@login_required
@role_required('admin')
def admin_dashboard():
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT COUNT(*) as n FROM users WHERE role='trainee'");    t = c.fetchone()['n']
    c.execute("SELECT COUNT(*) as n FROM users WHERE role='supervisor'"); s = c.fetchone()['n']
    c.execute("SELECT COUNT(*) as n FROM groups");                        g = c.fetchone()['n']
    c.execute("SELECT COUNT(*) as n FROM initiatives WHERE status='pending'"); p = c.fetchone()['n']
    c.execute("SELECT COUNT(*) as n FROM attendance WHERE date=%s", (str(date.today()),)); a = c.fetchone()['n']
    stats = {'trainees': t, 'supervisors': s, 'groups': g, 'pending': p, 'today_present': a}
    # ✅ IMPROVEMENT 2: load supervisors per group
    c.execute("SELECT g.*, "
              "(SELECT COUNT(*) FROM users WHERE group_id=g.id AND role='trainee') as trainee_count "
              "FROM groups g")
    groups_raw = c.fetchall()
    groups = []
    for grp in groups_raw:
        c.execute("""SELECT u.id, u.name FROM users u
                     JOIN group_supervisors gs ON u.id=gs.supervisor_id
                     WHERE gs.group_id=%s""", (grp['id'],))
        grp_dict = dict(grp)
        grp_dict['supervisors'] = c.fetchall()
        groups.append(grp_dict)
    c.execute("SELECT u.*, g.name as group_name FROM users u "
              "LEFT JOIN groups g ON u.group_id=g.id WHERE u.role != 'admin' ORDER BY u.role, u.name")
    users = c.fetchall()
    c.execute("SELECT i.*, u.name as trainee_name, g.name as group_name "
              "FROM initiatives i JOIN users u ON i.user_id=u.id "
              "LEFT JOIN groups g ON u.group_id=g.id ORDER BY i.uploaded_at DESC")
    initiatives = c.fetchall()
    c.execute("SELECT * FROM groups"); all_groups = c.fetchall()
    c.execute("SELECT * FROM users WHERE role='supervisor'")
    all_supervisors = c.fetchall()
    conn.close()
    return render_template('admin_dashboard.html', stats=stats, groups=groups,
                           users=users, initiatives=initiatives, all_groups=all_groups,
                           all_supervisors=all_supervisors, today=str(date.today()))

# ✅ IMPROVEMENT 1: create group with optional cancel (just redirect)
@app.route('/admin/group/create', methods=['POST'])
@login_required
@role_required('admin')
def admin_create_group():
    action = request.form.get('action', 'save')
    if action == 'cancel':
        return redirect(url_for('admin_dashboard'))
    name = request.form['name'].strip()
    if not name:
        flash('اسم المجموعة مطلوب', 'danger')
        return redirect(url_for('admin_dashboard'))
    sd = request.form.get('start_date') or str(date.today())
    conn = get_db(); c = conn.cursor()
    c.execute("INSERT INTO groups (name,start_date) VALUES (%s,%s) RETURNING id", (name, sd))
    gid = c.fetchone()['id']
    # Assign supervisor if selected
    sup = request.form.get('supervisor_id')
    if sup:
        c.execute("INSERT INTO group_supervisors (group_id,supervisor_id) VALUES (%s,%s) "
                  "ON CONFLICT DO NOTHING", (gid, sup))
    conn.commit(); conn.close()
    flash(f'تم إنشاء المجموعة "{name}" بنجاح', 'success')
    return redirect(url_for('admin_dashboard'))

# ✅ IMPROVEMENT 1: delete group
@app.route('/admin/group/<int:gid>/delete', methods=['POST'])
@login_required
@role_required('admin')
def admin_delete_group(gid):
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT name FROM groups WHERE id=%s", (gid,))
    grp = c.fetchone()
    if grp:
        c.execute("DELETE FROM group_supervisors WHERE group_id=%s", (gid,))
        c.execute("UPDATE users SET group_id=NULL WHERE group_id=%s", (gid,))
        c.execute("DELETE FROM groups WHERE id=%s", (gid,))
        conn.commit()
        flash(f'تم حذف المجموعة "{grp["name"]}"', 'success')
    conn.close()
    return redirect(url_for('admin_dashboard'))

# ✅ IMPROVEMENT 2: assign supervisor to group (supports multiple)
@app.route('/admin/group/<int:gid>/assign', methods=['POST'])
@login_required
@role_required('admin')
def admin_assign_supervisor(gid):
    supervisor_id = request.form['supervisor_id']
    conn = get_db(); c = conn.cursor()
    c.execute("INSERT INTO group_supervisors (group_id,supervisor_id) VALUES (%s,%s) ON CONFLICT DO NOTHING",
              (gid, supervisor_id))
    conn.commit(); conn.close()
    flash('تم إضافة المشرف للمجموعة بنجاح', 'success')
    return redirect(url_for('admin_dashboard'))

# ✅ IMPROVEMENT 2: remove supervisor from group
@app.route('/admin/group/<int:gid>/remove-supervisor/<int:sid>', methods=['POST'])
@login_required
@role_required('admin')
def admin_remove_supervisor(gid, sid):
    conn = get_db(); c = conn.cursor()
    c.execute("DELETE FROM group_supervisors WHERE group_id=%s AND supervisor_id=%s", (gid, sid))
    conn.commit(); conn.close()
    flash('تم إزالة المشرف من المجموعة', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/user/<int:uid>/toggle', methods=['POST'])
@login_required
@role_required('admin')
def admin_toggle_user(uid):
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT is_active FROM users WHERE id=%s", (uid,))
    user = c.fetchone()
    c.execute("UPDATE users SET is_active=%s WHERE id=%s", (0 if user['is_active'] else 1, uid))
    conn.commit(); conn.close()
    flash('تم تحديث حالة الحساب', 'success')
    return redirect(url_for('admin_dashboard'))

# ✅ IMPROVEMENT 6: delete user permanently
@app.route('/admin/user/<int:uid>/delete', methods=['POST'])
@login_required
@role_required('admin')
def admin_delete_user(uid):
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT name, role FROM users WHERE id=%s", (uid,))
    user = c.fetchone()
    if not user or user['role'] == 'admin':
        conn.close()
        flash('لا يمكن حذف هذا الحساب', 'danger')
        return redirect(url_for('admin_dashboard'))
    # Remove from group_supervisors if supervisor
    c.execute("DELETE FROM group_supervisors WHERE supervisor_id=%s", (uid,))
    # Delete attendance records
    c.execute("DELETE FROM attendance WHERE user_id=%s", (uid,))
    # Delete initiatives
    c.execute("SELECT filename FROM initiatives WHERE user_id=%s", (uid,))
    inits = c.fetchall()
    for ini in inits:
        if ini['filename']:
            delete_from_cloudinary(ini['filename'])
    c.execute("DELETE FROM initiatives WHERE user_id=%s", (uid,))
    # Delete user
    c.execute("DELETE FROM users WHERE id=%s", (uid,))
    conn.commit(); conn.close()
    flash(f'تم حذف المستخدم "{user["name"]}" نهائياً', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/user/<int:uid>/move', methods=['POST'])
@login_required
@role_required('admin')
def admin_move_user(uid):
    conn = get_db(); c = conn.cursor()
    c.execute("UPDATE users SET group_id=%s WHERE id=%s", (request.form['group_id'] or None, uid))
    conn.commit(); conn.close()
    flash('تم نقل المشترك بنجاح', 'success')
    return redirect(url_for('admin_dashboard'))

# ✅ IMPROVEMENT 4: invite codes only for supervisors
@app.route('/admin/invite', methods=['POST'])
@login_required
@role_required('admin')
def admin_create_invite():
    code = secrets.token_urlsafe(8)
    conn = get_db(); c = conn.cursor()
    c.execute("INSERT INTO invite_codes (code,role,created_by) VALUES (%s,'supervisor',%s)",
              (code, session['user_id']))
    conn.commit(); conn.close()
    flash(f'كود دعوة المشرف الجديد: {code}', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/export')
@login_required
@role_required('admin')
def admin_export():
    fmt = request.args.get('format', 'excel'); group_id = request.args.get('group_id')
    conn = get_db(); c = conn.cursor()
    if group_id:
        c.execute("SELECT name FROM groups WHERE id=%s", (group_id,))
        g = c.fetchone(); gname = g['name'] if g else 'Group'
        c.execute("SELECT u.name, a.date, a.status FROM attendance a "
                  "JOIN users u ON a.user_id=u.id WHERE u.group_id=%s ORDER BY a.date DESC", (group_id,))
    else:
        gname = 'All_Groups'
        c.execute("SELECT u.name, a.date, a.status FROM attendance a "
                  "JOIN users u ON a.user_id=u.id ORDER BY a.date DESC")
    records = c.fetchall(); conn.close()
    return export_attendance(records, gname, fmt)

if __name__ == '__main__':
    init_db()
    app.run(debug=True)